from flask import Flask, render_template, request, redirect
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

EXCEL_FILE = 'datos.xlsx'

def crear_excel_si_no_existe():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(['Nombre', 'Correo', 'Teléfono'])
        wb.save(EXCEL_FILE)

@app.route('/', methods=['GET', 'POST'])
def formulario():
    if request.method == 'POST':
        nombre = request.form.get('nombre')
        correo = request.form.get('correo')
        telefono = request.form.get('telefono')

        crear_excel_si_no_existe()

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append([nombre, correo, telefono])
        wb.save(EXCEL_FILE)

        return redirect('/gracias')
    return render_template('formulario.html')

@app.route('/gracias')
def gracias():
    return '¡Gracias por enviar tu información!'

if __name__ == '__main__':
    app.run(debug=True)